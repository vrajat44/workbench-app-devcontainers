"""
WB Data Catalog v2 — FastAPI backend.

Single-project scope: profiles the GCP project provided via settings or
defaults to the current workspace project.
"""

from __future__ import annotations

import logging
import os
from contextlib import asynccontextmanager
from pathlib import Path

if os.environ.get("LOG_FORMAT") == "json":
    logging.basicConfig(
        format='{"time":"%(asctime)s","level":"%(levelname)s","module":"%(module)s","message":"%(message)s"}',
        level=logging.INFO,
    )
else:
    logging.basicConfig(format="%(asctime)s %(levelname)s %(message)s", level=logging.INFO)
logger = logging.getLogger("datacatalog")
from typing import Any, Optional

from fastapi import BackgroundTasks, FastAPI, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles

from api_models import (
    CatalogResponse,
    ChartSuggestion,
    ChartsSuggestResponse,
    JobStartResponse,
    ProfileStatusResponse,
    TableSummary,
)
from bq_preview import MAX_PREVIEW_ROWS, preview_table
from chart_advisor import suggest_charts
from profiling_runner import (
    job_state,
    load_table_info,
    profile_status_from_gcs_and_jobs,
    run_semantic_profile_async,
    run_technical_profile_async,
)
from verily_profiler import discover_datasets, discover_tables, get_table_api_metadata, scan_profile_availability
from verily_profiler.storage import parse_fq_table, tech_object_path, sem_object_path, read_json_if_exists
from verily_profiler.llm import detect_available_model


import time as _time

_catalog_cache: dict[str, tuple[float, Any]] = {}
_CATALOG_CACHE_TTL = 120  # seconds
_terminology_cache: dict[str, tuple[float, Any]] = {}
_cohort_dims_cache: dict[str, tuple[float, Any]] = {}
_SLOW_CACHE_TTL = 600  # seconds — for GCS-heavy endpoints
_WS_CACHE_TTL = 300  # seconds — workspace dataset list
_scan_cache: dict[str, tuple[float, Any]] = {}


def _cached_scan() -> dict[str, dict[str, Any]]:
    """Cached wrapper around scan_profile_availability."""
    cache_key = f"{DATA_PROJECT}:{PROFILE_BUCKET}"
    now = _time.time()
    cached = _scan_cache.get(cache_key)
    if cached is not None:
        ts, resp = cached
        if now - ts < _SLOW_CACHE_TTL:
            return resp
    result = scan_profile_availability(PROFILE_BUCKET, DATA_PROJECT, billing_project_id=BILLING_PROJECT)
    _scan_cache[cache_key] = (_time.time(), result)
    return result


def _profiling_for_catalog_row(fq: str, prof: dict[str, bool]) -> dict[str, str]:
    """Merge GCS profile presence with in-memory job running flags."""
    flags = job_state.running_flags(fq)
    tech = "running" if flags["technical"] else ("available" if prof["technical"] else "none")
    sem = "running" if flags["semantic"] else ("available" if prof["semantic"] else "none")
    return {"technical": tech, "semantic": sem}


# ── Settings ──────────────────────────────────────────────────────────────────

DEFAULT_BILLING_PROJECT = "wb-shrewd-papaya-8403"

PROFILE_BUCKET = ""
BILLING_PROJECT: str = ""
DATA_PROJECT: str = ""
GEMINI_MODEL: Optional[str] = None      # for profiling (tech + semantic)
GEMINI_LOCATION: Optional[str] = None   # Vertex AI location (us-central1, global, etc.)
CHAT_MODEL: Optional[str] = None         # for chat (None = use auto-detected)
PROJECT_DISPLAY_NAME: str = ""
FRONTEND_DIST = Path(
    os.environ.get("FRONTEND_DIST", str(Path(__file__).resolve().parent / "static")),
)


def _derive_bucket(project_id: str) -> str:
    return f"metadata-json-{project_id}" if project_id else ""


_ws_name_cache: dict[str, str] = {}


def _resolve_project_name(gcp_project_id: str) -> str:
    """Look up workspace display name from wb CLI by matching googleProjectId."""
    if gcp_project_id in _ws_name_cache:
        return _ws_name_cache[gcp_project_id]
    import subprocess, json as _json
    try:
        import shutil
        wb_path = shutil.which("wb")
        if not wb_path:
            return None
        result = subprocess.run(
            [wb_path, "workspace", "list", "--format=json"],
            capture_output=True, text=True, timeout=60,
        )
        if result.returncode != 0:
            return ""
        for ws in _json.loads(result.stdout):
            gcp = ws.get("googleProjectId", "")
            name = ws.get("name", "")
            if gcp:
                _ws_name_cache[gcp] = name
        return _ws_name_cache.get(gcp_project_id, "")
    except Exception:
        pass
    return ""


def _detect_current_workspace_project() -> str:
    """Get the GCP project of the workspace where this app is running."""
    data = _run_wb(["workspace", "describe", "--format=json"], timeout=15)
    if isinstance(data, dict):
        return data.get("googleProjectId", "")
    return ""


def _ensure_catalog_context_exists():
    import threading
    from verily_profiler.storage import read_catalog_context, regenerate_catalog_context

    def _work():
        try:
            existing = read_catalog_context(PROFILE_BUCKET, DATA_PROJECT, billing_project_id=BILLING_PROJECT)
            if existing:
                logger.info(f"Catalog context exists ({len(existing)} chars)")
                return
            avail = _cached_scan()
            profiled = [fq for fq, info in avail.items() if info.get("technical") or info.get("semantic")]
            if not profiled:
                return
            logger.info(f"Generating catalog context for {len(profiled)} existing profiles...")
            regenerate_catalog_context(PROFILE_BUCKET, DATA_PROJECT, billing_project_id=BILLING_PROJECT)
        except Exception as e:
            logger.warning(f"Startup context generation failed: {e}")

    threading.Thread(target=_work, daemon=True).start()


def _prewarm_caches():
    import threading

    def _work():
        global PROJECT_DISPLAY_NAME

        _load_workspaces()

        if DATA_PROJECT:
            name = _resolve_project_name(DATA_PROJECT)
            if name:
                PROJECT_DISPLAY_NAME = name
                logger.info(f"Project: {name} ({DATA_PROJECT})")

        import urllib.request
        base = "http://127.0.0.1:8080"
        for path in ["/api/catalog", "/api/terminology", "/api/cohorts/dimensions"]:
            try:
                urllib.request.urlopen(f"{base}{path}", timeout=15)
                logger.info(f"Cache warm: {path}")
            except Exception as e:
                logger.warning(f"Cache warm failed {path}: {e}")
                break

    threading.Thread(target=_work, daemon=True).start()


@asynccontextmanager
async def lifespan(app: FastAPI):
    global BILLING_PROJECT, DATA_PROJECT, GEMINI_MODEL, GEMINI_LOCATION, CHAT_MODEL, PROFILE_BUCKET, PROJECT_DISPLAY_NAME
    BILLING_PROJECT = os.environ.get("GCP_PROJECT_ID") or os.environ.get("BILLING_PROJECT_ID", "")
    if not BILLING_PROJECT:
        BILLING_PROJECT = _detect_current_workspace_project()
        if BILLING_PROJECT:
            logger.info(f"Billing project auto-detected from current workspace: {BILLING_PROJECT}")
    if not BILLING_PROJECT:
        BILLING_PROJECT = DEFAULT_BILLING_PROJECT
        logger.info(f"Using default billing project: {BILLING_PROJECT}")
    DATA_PROJECT = (os.environ.get("DATA_PROJECT_ID") or "").strip()
    if not DATA_PROJECT:
        DATA_PROJECT = BILLING_PROJECT
        logger.warning("DATA_PROJECT_ID not set — defaulting to billing project %s", BILLING_PROJECT)
    GEMINI_MODEL = os.environ.get("GEMINI_MODEL") or None
    GEMINI_LOCATION = os.environ.get("GEMINI_LOCATION") or None
    from verily_profiler.llm import set_model_settings
    set_model_settings(GEMINI_MODEL, GEMINI_LOCATION)
    CHAT_MODEL = os.environ.get("CHAT_MODEL") or None
    PROFILE_BUCKET = _derive_bucket(BILLING_PROJECT)
    if not BILLING_PROJECT:
        logger.info("GCP_PROJECT_ID not set — configure via UI Settings")
    elif PROFILE_BUCKET:
        result = _ensure_bucket(PROFILE_BUCKET, BILLING_PROJECT)
        logger.info(f"Bucket: {result}")
        if DATA_PROJECT:
            _ensure_catalog_context_exists()
            _prewarm_caches()
    yield


app = FastAPI(title="WB Data Catalog v2", lifespan=lifespan)
_CORS_ORIGINS = os.environ.get("CORS_ORIGINS", "http://localhost:5173").split(",")
app.add_middleware(
    CORSMiddleware,
    allow_origins=_CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


def _fq(project_id: str, dataset_id: str, table_id: str) -> str:
    return f"{project_id}.{dataset_id}.{table_id}"


@app.get("/api/health")
def health():
    return {"status": "ok", "data_project": DATA_PROJECT, "profile_bucket": PROFILE_BUCKET}


@app.get("/api/health/deep")
def health_deep():
    """Readiness probe: validates GCS bucket access."""
    issues: list[str] = []
    if not BILLING_PROJECT:
        issues.append("billing_project not configured")
    if PROFILE_BUCKET:
        try:
            from google.cloud import storage
            c = storage.Client(project=BILLING_PROJECT)
            c.bucket(PROFILE_BUCKET.replace("gs://", "")).exists()
        except Exception as e:
            issues.append(f"GCS bucket unreachable: {e}")
    if issues:
        raise HTTPException(503, {"status": "degraded", "issues": issues})
    return {"status": "ok", "data_project": DATA_PROJECT, "profile_bucket": PROFILE_BUCKET}


# ── Workspace discovery via wb CLI ──────────────────────────────────────────

_workspaces_cache: dict[str, tuple[float, Any]] = {}
_WS_DISK_CACHE = Path(__file__).resolve().parent / ".wb_workspaces_cache.json"
_WS_DISK_TTL = 86400  # 24 hours


class WbAuthError(Exception):
    pass


def _run_wb(args: list[str], timeout: int = 60) -> list | dict | None:
    import subprocess, json as _json, shutil
    wb_path = shutil.which("wb")
    if not wb_path:
        return None
    try:
        result = subprocess.run(
            [wb_path] + args,
            capture_output=True, text=True, timeout=timeout,
        )
        if result.returncode != 0:
            combined = (result.stderr or "") + (result.stdout or "")
            if "open the following address" in combined or "auth" in combined.lower():
                raise WbAuthError("Workbench CLI session expired. Run `wb auth login` in your terminal to re-authenticate.")
            return None
        return _json.loads(result.stdout)
    except WbAuthError:
        raise
    except Exception:
        return None


def _load_workspaces() -> dict[str, Any]:
    """Load workspaces from memory cache → disk cache → wb CLI (background refresh)."""
    import json as _json

    mem = _workspaces_cache.get("all")
    if mem is not None:
        return mem[1]

    if _WS_DISK_CACHE.is_file():
        try:
            disk = _json.loads(_WS_DISK_CACHE.read_text())
            ts = disk.get("_cached_at", 0)
            response = {"workspaces": disk.get("workspaces", [])}
            _workspaces_cache["all"] = (ts, response)
            if _time.time() - ts > _WS_DISK_TTL:
                _refresh_workspaces_background()
            return response
        except Exception:
            pass

    _refresh_workspaces_background()
    return {"workspaces": []}


def _refresh_workspaces_background():
    """Fetch workspace list from wb CLI in background and write to disk + memory."""
    import threading, json as _json

    def _work():
        data = _run_wb(["workspace", "list", "--format=json"], timeout=600)
        if data is None:
            return
        workspaces = []
        for ws in data:
            workspaces.append({
                "id": ws.get("id", ""),
                "name": ws.get("name", ""),
                "gcp_project": ws.get("googleProjectId", ""),
                "role": ws.get("highestRole", ""),
            })
        workspaces.sort(key=lambda w: w["name"].lower())
        response = {"workspaces": workspaces}
        now = _time.time()
        _workspaces_cache["all"] = (now, response)
        try:
            _WS_DISK_CACHE.write_text(_json.dumps({"_cached_at": now, "workspaces": workspaces}, indent=2))
            logger.info(f"Workspace cache refreshed: {len(workspaces)} workspaces")
        except Exception as e:
            logger.warning(f"Failed to write workspace cache: {e}")

    threading.Thread(target=_work, daemon=True).start()


@app.get("/api/workspaces")
def api_workspaces():
    """List all workspaces the user has access to (from cache)."""
    return _load_workspaces()


@app.get("/api/workspaces/{workspace_id}/datasets")
def api_workspace_datasets(workspace_id: str):
    """List BQ datasets in a specific workspace."""
    import re
    if not re.match(r"^[a-zA-Z0-9_-]+$", workspace_id):
        raise HTTPException(400, "Invalid workspace ID")

    cache_key = workspace_id
    now = _time.time()
    cached = _workspaces_cache.get(cache_key)
    if cached is not None:
        ts, resp = cached
        if now - ts < _WS_CACHE_TTL:
            return resp

    try:
        data = _run_wb(["resource", "list", f"--workspace={workspace_id}", "--format=json"])
    except WbAuthError as e:
        raise HTTPException(401, str(e))
    if data is None:
        raise HTTPException(503, "wb CLI unavailable or timed out")

    datasets = []
    for r in data:
        if r.get("resourceType") != "BQ_DATASET":
            continue
        datasets.append({
            "id": r.get("id", ""),
            "project_id": r.get("projectId", ""),
            "dataset_id": r.get("datasetId", ""),
            "num_tables": r.get("numTables"),
            "type": r.get("stewardshipType", ""),
            "location": r.get("location", ""),
        })
    datasets.sort(key=lambda d: d["id"])
    response = {"datasets": datasets}
    _workspaces_cache[cache_key] = (_time.time(), response)
    return response


@app.get("/api/config")
def api_config():
    return {
        "billing_project": BILLING_PROJECT,
        "data_project": DATA_PROJECT,
        "data_project_name": PROJECT_DISPLAY_NAME,
        "profile_bucket": PROFILE_BUCKET,
        "gemini_model": GEMINI_MODEL,
        "gemini_location": GEMINI_LOCATION,
        "configured": bool(BILLING_PROJECT),
        "default_billing_project": DEFAULT_BILLING_PROJECT,
    }


SUPPORTED_LOCATIONS = ["auto", "us-central1", "global", "europe-west4"]

SUPPORTED_MODELS = [
    {"id": "", "label": "Auto-detect (recommended)"},
    {"id": "gemini-2.5-flash", "label": "Gemini 2.5 Flash"},
    {"id": "gemini-2.5-pro", "label": "Gemini 2.5 Pro"},
    {"id": "gemini-3.5-flash", "label": "Gemini 3.5 Flash"},
    {"id": "gemini-3.5-pro", "label": "Gemini 3.5 Pro"},
]


@app.get("/api/models")
def api_models():
    from verily_profiler.llm import _resolved_model, _resolved_location
    return {
        "models": SUPPORTED_MODELS,
        "locations": SUPPORTED_LOCATIONS,
        "detected_model": _resolved_model,
        "detected_location": _resolved_location,
    }


@app.post("/api/models/test")
def api_test_model(body: dict[str, Any]):
    """Test if a specific model + location combination works."""
    model = body.get("model", "")
    location = body.get("location", "us-central1")
    if not model:
        from verily_profiler.llm import detect_available_model
        try:
            detected = detect_available_model(BILLING_PROJECT)
            from verily_profiler.llm import _resolved_location
            return {"status": "ok", "model": detected, "location": _resolved_location}
        except Exception as e:
            return {"status": "error", "error": str(e)}
    try:
        from verily_profiler.llm import _get_client
        from google.genai.types import GenerateContentConfig
        client = _get_client(BILLING_PROJECT, location)
        client.models.generate_content(
            model=model,
            contents="Reply with just the word: ok",
            config=GenerateContentConfig(temperature=0.0, max_output_tokens=8),
        )
        return {"status": "ok", "model": model, "location": location}
    except Exception as e:
        return {"status": "error", "model": model, "location": location, "error": str(e)}


def _ensure_bucket(bucket_name: str, project_id: str) -> dict[str, Any]:
    """Check if the profile bucket exists; create it if it doesn't."""
    from google.cloud import storage

    client = storage.Client(project=project_id)
    bucket_ref = client.bucket(bucket_name)
    try:
        bucket_ref.reload(client=client)
        return {"bucket": bucket_name, "action": "exists"}
    except Exception:
        try:
            bucket_ref.storage_class = "STANDARD"
            client.create_bucket(bucket_ref, project=project_id, location="us")
            logger.info(f"Created bucket: {bucket_name}")
            return {"bucket": bucket_name, "action": "created"}
        except Exception as e:
            return {"bucket": bucket_name, "action": "error", "error": str(e)}


@app.put("/api/settings")
def api_update_settings(body: dict[str, Any]):
    """
    Update runtime settings from the UI.
    Accepts: { billing_project?, data_project?, gemini_model? }
    """
    global BILLING_PROJECT, DATA_PROJECT, GEMINI_MODEL, GEMINI_LOCATION, PROFILE_BUCKET, PROJECT_DISPLAY_NAME
    if "billing_project" in body:
        BILLING_PROJECT = str(body["billing_project"]).strip()
        PROFILE_BUCKET = _derive_bucket(BILLING_PROJECT)
    if "data_project" in body:
        DATA_PROJECT = str(body["data_project"]).strip() or BILLING_PROJECT
        PROJECT_DISPLAY_NAME = _resolve_project_name(DATA_PROJECT) if DATA_PROJECT else ""
    elif not DATA_PROJECT and BILLING_PROJECT:
        DATA_PROJECT = BILLING_PROJECT
    if "gemini_model" in body:
        val = str(body["gemini_model"]).strip()
        GEMINI_MODEL = val if val else None
    if "gemini_location" in body:
        val = str(body["gemini_location"]).strip()
        GEMINI_LOCATION = val if val and val != "auto" else None
    from verily_profiler.llm import set_model_settings
    set_model_settings(GEMINI_MODEL, GEMINI_LOCATION)
    from chat_handler import invalidate_context_cache, chat_store
    invalidate_context_cache()
    _invalidate_profiling_caches()
    chat_store._sessions.clear()
    logger.info("Settings changed: all caches invalidated, chat sessions cleared")

    bucket_status: dict[str, Any] = {}
    if BILLING_PROJECT and PROFILE_BUCKET:
        bucket_status = _ensure_bucket(PROFILE_BUCKET, BILLING_PROJECT)
        logger.info(f"Bucket check: {bucket_status}")
        _ensure_catalog_context_exists()

    cfg = api_config()
    cfg["bucket_status"] = bucket_status
    return cfg


# ── Discovery & catalog ───────────────────────────────────────────────────────

@app.get("/api/datasets")
def api_datasets():
    if not DATA_PROJECT:
        raise HTTPException(503, "No data project configured")
    return {"project_id": DATA_PROJECT, "datasets": discover_datasets(DATA_PROJECT, billing_project=BILLING_PROJECT)}


@app.get("/api/datasets/{dataset_id}/tables")
def api_dataset_tables(dataset_id: str):
    if not DATA_PROJECT:
        raise HTTPException(503, "No data project configured")
    tables = discover_tables(DATA_PROJECT, dataset_id, billing_project=BILLING_PROJECT)
    profile_index: dict[str, dict[str, bool]] = {}
    if PROFILE_BUCKET:
        try:
            profile_index = _cached_scan()
        except Exception as e:
            logger.warning(f"Profile scan failed: {e}")
    out = []
    for t in tables:
        fq = t.fq_name
        prof = profile_index.get(fq, {"technical": False, "semantic": False})
        ts = getattr(t, "creation_time", None)
        out.append(
            TableSummary(
                fq_table=fq,
                project_id=t.project_id,
                dataset_id=t.dataset_id,
                table_id=t.table_id,
                row_count=t.row_count,
                size_bytes=t.size_bytes,
                table_type=t.table_type,
                column_count=len(t.columns),
                creation_time=str(ts) if ts else None,
                profiling=_profiling_for_catalog_row(fq, prof),
                business_name=prof.get("business_name"),
                table_definition=prof.get("table_definition"),
            ).model_dump()
        )
    return {"dataset_id": dataset_id, "tables": out}


@app.get("/api/catalog")
def api_catalog(refresh: bool = False):
    """All datasets with table summaries and profiling flags."""
    from concurrent.futures import ThreadPoolExecutor, as_completed

    if not DATA_PROJECT:
        raise HTTPException(503, "No data project configured")

    cache_key = f"{DATA_PROJECT}:{PROFILE_BUCKET}"
    if refresh:
        _catalog_cache.pop(cache_key, None)
        _scan_cache.pop(cache_key, None)

    now = _time.time()
    cached_entry = _catalog_cache.get(cache_key)
    if cached_entry is not None:
        ts, cached_resp = cached_entry
        if now - ts < _CATALOG_CACHE_TTL:
            return cached_resp

    # Run GCS profile scan and BQ dataset discovery in parallel
    profile_index: dict[str, dict[str, bool]] = {}
    datasets: list[str] = []

    def _scan_profiles():
        if not PROFILE_BUCKET:
            return {}
        try:
            return _cached_scan()
        except Exception as e:
            logger.warning(f"Profile scan failed: {e}")
            return {}

    def _list_datasets():
        return discover_datasets(DATA_PROJECT, billing_project=BILLING_PROJECT)

    with ThreadPoolExecutor(max_workers=2) as ex:
        f_profiles = ex.submit(_scan_profiles)
        f_datasets = ex.submit(_list_datasets)

        profile_index = f_profiles.result()

        try:
            datasets = f_datasets.result()
        except Exception as e:
            logger.warning(f"Dataset discovery failed: {e}")

    # Discover tables across all datasets in parallel
    ds_tables: dict[str, list] = {}

    def _discover_ds(ds_id: str):
        return ds_id, discover_tables(DATA_PROJECT, ds_id, billing_project=BILLING_PROJECT)

    with ThreadPoolExecutor(max_workers=min(8, max(len(datasets), 1))) as ex:
        futures = {ex.submit(_discover_ds, ds): ds for ds in datasets}
        for f in as_completed(futures):
            try:
                ds_id, tables = f.result()
                ds_tables[ds_id] = tables
            except Exception as e:
                ds_id = futures[f]
                logger.warning(f"Table discovery failed for {ds_id}: {e}")
                ds_tables[ds_id] = []

    # Pre-fetch tech profiles in parallel for tables that need them
    all_tables_flat = [(ds, t) for ds in datasets for t in ds_tables.get(ds, [])]
    needs_tech = [
        t.fq_name for _, t in all_tables_flat
        if (t.row_count is None or t.size_bytes is None)
        and profile_index.get(t.fq_name, {}).get("technical")
    ]
    tech_cache: dict[str, dict | None] = {}
    if needs_tech:
        def _load_tech(fq: str):
            p_, d_, t_ = parse_fq_table(fq)
            return fq, read_json_if_exists(PROFILE_BUCKET, tech_object_path(p_, d_, t_), BILLING_PROJECT)
        with ThreadPoolExecutor(max_workers=10) as ex:
            for fq, data in ex.map(lambda f: _load_tech(f), needs_tech):
                tech_cache[fq] = data

    result: list[dict[str, Any]] = []
    for ds in datasets:
        tables = ds_tables.get(ds, [])
        rows = []
        for t in tables:
            fq = t.fq_name
            prof = profile_index.get(fq, {"technical": False, "semantic": False})
            ts = getattr(t, "creation_time", None)

            row_count = t.row_count
            size_bytes = t.size_bytes
            col_count = len(t.columns)

            tech_data = tech_cache.get(fq)
            if tech_data:
                if row_count is None and tech_data.get("row_count") is not None:
                    row_count = tech_data["row_count"]
                if size_bytes is None and tech_data.get("size_bytes") is not None:
                    size_bytes = tech_data["size_bytes"]
                if col_count == 0 and tech_data.get("columns"):
                    col_count = len(tech_data["columns"])

            rows.append(
                TableSummary(
                    fq_table=fq,
                    project_id=t.project_id,
                    dataset_id=t.dataset_id,
                    table_id=t.table_id,
                    row_count=row_count,
                    size_bytes=size_bytes,
                    table_type=t.table_type,
                    column_count=col_count,
                    creation_time=str(ts) if ts else None,
                    profiling=_profiling_for_catalog_row(fq, prof),
                    business_name=prof.get("business_name"),
                    table_definition=prof.get("table_definition"),
                ).model_dump()
            )
        result.append({"dataset_id": ds, "tables": rows})
    response = CatalogResponse(
        project_id=DATA_PROJECT,
        profile_bucket=PROFILE_BUCKET or "",
        datasets=result,
    ).model_dump()
    _catalog_cache[cache_key] = (_time.time(), response)
    return response


# ── Table-level endpoints ─────────────────────────────────────────────────────

@app.get("/api/projects/{project_id}/datasets/{dataset_id}/tables/{table_id}/preview")
def api_preview(project_id: str, dataset_id: str, table_id: str, limit: int = MAX_PREVIEW_ROWS):
    info = load_table_info(_fq(project_id, dataset_id, table_id), BILLING_PROJECT, project_id)
    if not info:
        raise HTTPException(404, "Table not found")
    return preview_table(info, BILLING_PROJECT, limit=limit)


@app.get("/api/projects/{project_id}/datasets/{dataset_id}/tables/{table_id}/profile/status")
def api_profile_status(project_id: str, dataset_id: str, table_id: str):
    fq = _fq(project_id, dataset_id, table_id)
    if not PROFILE_BUCKET:
        raise HTTPException(503, "PROFILE_GCS_BUCKET not configured")
    st = profile_status_from_gcs_and_jobs(fq, PROFILE_BUCKET, BILLING_PROJECT)
    return ProfileStatusResponse(**st).model_dump()


@app.get("/api/projects/{project_id}/datasets/{dataset_id}/tables/{table_id}/profile/technical")
def api_get_technical(project_id: str, dataset_id: str, table_id: str):
    fq = _fq(project_id, dataset_id, table_id)
    if not PROFILE_BUCKET:
        raise HTTPException(503, "PROFILE_GCS_BUCKET not configured")
    p, d, t = parse_fq_table(fq)
    data = read_json_if_exists(PROFILE_BUCKET, tech_object_path(p, d, t), BILLING_PROJECT)
    if not data:
        raise HTTPException(404, "Technical profile not found")
    return data


@app.get("/api/projects/{project_id}/datasets/{dataset_id}/tables/{table_id}/profile/semantic")
def api_get_semantic(project_id: str, dataset_id: str, table_id: str):
    fq = _fq(project_id, dataset_id, table_id)
    if not PROFILE_BUCKET:
        raise HTTPException(503, "PROFILE_GCS_BUCKET not configured")
    p, d, t = parse_fq_table(fq)
    data = read_json_if_exists(PROFILE_BUCKET, sem_object_path(p, d, t), BILLING_PROJECT)
    if not data:
        raise HTTPException(404, "Semantic profile not found")
    return data


EDITABLE_COL_FIELDS = {"definition", "sensitivity", "confidence", "unit_of_measure", "measurement_method"}


@app.put("/api/projects/{project_id}/datasets/{dataset_id}/tables/{table_id}/profile/semantic")
def api_update_semantic(project_id: str, dataset_id: str, table_id: str, body: dict[str, Any]):
    """Merge user edits into the existing semantic profile and save to GCS."""
    fq = _fq(project_id, dataset_id, table_id)
    if not PROFILE_BUCKET:
        raise HTTPException(503, "PROFILE_GCS_BUCKET not configured")
    from verily_profiler.storage import upload_json

    p, d, t = parse_fq_table(fq)
    profile = read_json_if_exists(PROFILE_BUCKET, sem_object_path(p, d, t), BILLING_PROJECT)
    if not profile:
        raise HTTPException(404, "Semantic profile not found")

    edits_by_name = {c["name"]: c for c in body.get("columns", []) if isinstance(c, dict) and "name" in c}
    for col in profile.get("columns", []):
        edit = edits_by_name.get(col.get("name"))
        if not edit:
            continue
        for field in EDITABLE_COL_FIELDS:
            if field in edit:
                col[field] = edit[field]

    upload_json(PROFILE_BUCKET, sem_object_path(p, d, t), profile, BILLING_PROJECT)
    return profile


def _invalidate_profiling_caches():
    """Clear all caches that depend on profiling results."""
    _scan_cache.clear()
    _catalog_cache.clear()
    _cohort_dims_cache.clear()
    _terminology_cache.clear()
    _terminology_slim_cache.clear()
    _col_values_cache.clear()


@app.post("/api/projects/{project_id}/datasets/{dataset_id}/tables/{table_id}/profile/technical")
async def api_run_technical(
    project_id: str,
    dataset_id: str,
    table_id: str,
    background_tasks: BackgroundTasks,
):
    fq = _fq(project_id, dataset_id, table_id)
    if not PROFILE_BUCKET:
        raise HTTPException(503, "PROFILE_GCS_BUCKET not configured")
    _ensure_bucket(PROFILE_BUCKET, BILLING_PROJECT)
    jid, started = job_state.try_start(fq, "technical")
    if not started:
        return JobStartResponse(job_id=jid, status="running").model_dump()

    async def _job():
        try:
            await run_technical_profile_async(
                fq_table=fq,
                bucket=PROFILE_BUCKET,
                billing_project=BILLING_PROJECT,
                data_project=project_id,
                job_id=jid,
            )
        finally:
            _invalidate_profiling_caches()

    background_tasks.add_task(_job)
    return JobStartResponse(job_id=jid, status="running").model_dump()


@app.post("/api/projects/{project_id}/datasets/{dataset_id}/tables/{table_id}/profile/semantic")
async def api_run_semantic(
    project_id: str,
    dataset_id: str,
    table_id: str,
    background_tasks: BackgroundTasks,
):
    fq = _fq(project_id, dataset_id, table_id)
    if not PROFILE_BUCKET:
        raise HTTPException(503, "PROFILE_GCS_BUCKET not configured")
    _ensure_bucket(PROFILE_BUCKET, BILLING_PROJECT)
    p, d, t = parse_fq_table(fq)
    if not read_json_if_exists(PROFILE_BUCKET, tech_object_path(p, d, t), BILLING_PROJECT):
        raise HTTPException(409, "Run technical profiling first")
    jid, started = job_state.try_start(fq, "semantic")
    if not started:
        return JobStartResponse(job_id=jid, status="running").model_dump()

    async def _job():
        try:
            await run_semantic_profile_async(
                fq_table=fq,
                bucket=PROFILE_BUCKET,
                billing_project=BILLING_PROJECT,
                data_project=project_id,
                model_name=GEMINI_MODEL,
                job_id=jid,
            )
        finally:
            _invalidate_profiling_caches()

    background_tasks.add_task(_job)
    return JobStartResponse(job_id=jid, status="running").model_dump()


# ── Charts / Explore ──────────────────────────────────────────────────────────

@app.post("/api/charts/suggest")
def api_charts_suggest(body: dict[str, Any]):
    technical = body.get("technical") or {}
    semantic = body.get("semantic")
    if not technical.get("columns"):
        raise HTTPException(400, "technical profile with columns is required")
    model = GEMINI_MODEL or detect_available_model(BILLING_PROJECT)
    charts_raw = suggest_charts(technical, semantic, model, BILLING_PROJECT)
    charts = [ChartSuggestion.model_validate(c) for c in charts_raw]
    return ChartsSuggestResponse(charts=charts).model_dump()


@app.post("/api/gw/compute/{project_id}/{dataset_id}/{table_id}")
def api_gw_compute(project_id: str, dataset_id: str, table_id: str, body: dict[str, Any]):
    """Execute a Graphic Walker computation query against BigQuery."""
    from gw_computation import execute_workflow

    fq = _fq(project_id, dataset_id, table_id)
    try:
        rows = execute_workflow(fq, body, billing_project=BILLING_PROJECT)
        return rows
    except Exception as e:
        raise HTTPException(400, f"Computation failed: {e}")


@app.get("/api/jobs/{job_id}")
def api_job(job_id: str):
    j = job_state.get_job(job_id)
    if not j:
        raise HTTPException(404, "Job not found")
    return j


# ── Bulk profiling ───────────────────────────────────────────────────────────

from bulk_profiler import bulk_manager


@app.post("/api/bulk-profile")
def api_bulk_profile(body: dict[str, Any]):
    """
    Start bulk profiling.
    Body: {
        tables: ["fq1", ...],
        mode: "technical" | "semantic" | "both",
        force?: boolean,
        terminology_domains?: ["clinical", "life_sciences"],
        doc_context?: str,
        run_join_inference?: boolean,
    }
    """
    tables = body.get("tables", [])
    mode = body.get("mode", "both")
    force = bool(body.get("force", False))
    terminology_domains = body.get("terminology_domains", [])
    doc_context = body.get("doc_context", "")
    run_join_inference = bool(body.get("run_join_inference", False))
    if not tables:
        raise HTTPException(400, "tables list is required")
    if mode not in ("technical", "semantic", "both"):
        mode = "both"
    if not DATA_PROJECT or not PROFILE_BUCKET:
        raise HTTPException(503, "Project or bucket not configured")
    _ensure_bucket(PROFILE_BUCKET, BILLING_PROJECT)

    batch_id = bulk_manager.start_batch(
        tables=tables,
        mode=mode,
        bucket=PROFILE_BUCKET,
        billing_project=BILLING_PROJECT,
        data_project=DATA_PROJECT,
        model_name=GEMINI_MODEL,
        force=force,
        terminology_domains=terminology_domains or None,
        doc_context=doc_context,
        run_join_inference=run_join_inference,
    )
    return {"batch_id": batch_id, "total": len(tables), "mode": mode, "force": force}


@app.get("/api/bulk-profile/{batch_id}")
def api_bulk_status(batch_id: str):
    """Get bulk profiling batch status."""
    batch = bulk_manager.get_batch(batch_id)
    if not batch:
        raise HTTPException(404, "Batch not found")
    return batch.summary()


# ── Terminology domains ───────────────────────────────────────────────────────

@app.get("/api/terminology-domains")
def api_terminology_domains():
    """Return available terminology domain presets for the Profiling Wizard."""
    from verily_profiler.terminology_domains import get_domains_metadata
    return {"domains": get_domains_metadata()}


# ── Profiling doc upload ──────────────────────────────────────────────────────

_ALLOWED_DOC_EXTENSIONS = {".pdf", ".md", ".txt", ".csv", ".xlsx"}
_USER_DOCS_PREFIX = "profiling/{data_project}/_user_docs/"


def _docs_gcs_prefix() -> str:
    return f"profiling/{DATA_PROJECT}/_user_docs/"


def _extract_text_from_upload(filename: str, content: bytes) -> str:
    """Extract text content from uploaded file based on extension."""
    ext = Path(filename).suffix.lower()
    if ext in (".md", ".txt"):
        return content.decode("utf-8", errors="replace")
    elif ext == ".csv":
        import csv
        import io
        text = content.decode("utf-8", errors="replace")
        reader = csv.DictReader(io.StringIO(text))
        lines = []
        for row in reader:
            pairs = [f"{k}: {v}" for k, v in row.items() if v]
            if pairs:
                lines.append("; ".join(pairs))
        return "\n".join(lines)
    elif ext == ".xlsx":
        import io
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [str(h or "") for h in rows[0]]
                for row in rows[1:]:
                    pairs = [f"{headers[i]}: {row[i]}" for i in range(len(headers)) if i < len(row) and row[i]]
                    if pairs:
                        lines.append("; ".join(pairs))
            return "\n".join(lines)
        except ImportError:
            return "[XLSX parsing requires openpyxl — install it to enable Excel uploads]"
    elif ext == ".pdf":
        import io
        try:
            import pdfplumber
            pages_text = []
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        pages_text.append(text)
            return "\n\n".join(pages_text)
        except ImportError:
            return "[PDF parsing requires pdfplumber — install it to enable PDF uploads]"
    return content.decode("utf-8", errors="replace")


@app.post("/api/profiling/docs")
async def api_upload_doc(file: UploadFile):
    """Upload a document to provide context for semantic profiling.

    Accepts PDF, MD, TXT, CSV, XLSX. Extracts text and stores in GCS.
    """
    if not DATA_PROJECT or not PROFILE_BUCKET:
        raise HTTPException(503, "Project or bucket not configured")

    filename = file.filename or "unnamed"
    ext = Path(filename).suffix.lower()
    if ext not in _ALLOWED_DOC_EXTENSIONS:
        raise HTTPException(
            400,
            f"Unsupported file type: {ext}. Allowed: {', '.join(sorted(_ALLOWED_DOC_EXTENSIONS))}",
        )

    content = await file.read()
    if len(content) > 10 * 1024 * 1024:  # 10 MB limit
        raise HTTPException(400, "File too large (max 10 MB)")

    extracted = _extract_text_from_upload(filename, content)

    # Store to GCS
    from google.cloud import storage as gcs_storage
    client = gcs_storage.Client(project=BILLING_PROJECT)
    bucket_obj = client.bucket(PROFILE_BUCKET.replace("gs://", ""))
    blob_path = _docs_gcs_prefix() + filename
    blob = bucket_obj.blob(blob_path)
    blob.upload_from_string(extracted, content_type="text/plain")

    preview = extracted[:500] if extracted else ""
    return {
        "filename": filename,
        "size": len(content),
        "extracted_chars": len(extracted),
        "preview": preview,
        "gcs_path": f"gs://{PROFILE_BUCKET}/{blob_path}",
    }


@app.get("/api/profiling/docs")
def api_list_docs():
    """List uploaded profiling context documents from GCS."""
    if not DATA_PROJECT or not PROFILE_BUCKET:
        raise HTTPException(503, "Project or bucket not configured")

    from google.cloud import storage as gcs_storage
    client = gcs_storage.Client(project=BILLING_PROJECT)
    bucket_obj = client.bucket(PROFILE_BUCKET.replace("gs://", ""))
    prefix = _docs_gcs_prefix()

    docs = []
    for blob in bucket_obj.list_blobs(prefix=prefix):
        name = blob.name.removeprefix(prefix)
        if not name:
            continue
        docs.append({
            "filename": name,
            "size": blob.size,
            "updated": blob.updated.isoformat() if blob.updated else "",
            "gcs_path": f"gs://{PROFILE_BUCKET}/{blob.name}",
        })
    return {"docs": docs}


@app.delete("/api/profiling/docs/{filename}")
def api_delete_doc(filename: str):
    """Delete an uploaded profiling context document from GCS."""
    if not DATA_PROJECT or not PROFILE_BUCKET:
        raise HTTPException(503, "Project or bucket not configured")

    from google.cloud import storage as gcs_storage
    client = gcs_storage.Client(project=BILLING_PROJECT)
    bucket_obj = client.bucket(PROFILE_BUCKET.replace("gs://", ""))
    blob_path = _docs_gcs_prefix() + filename
    blob = bucket_obj.blob(blob_path)

    if not blob.exists():
        raise HTTPException(404, f"Document not found: {filename}")

    blob.delete()
    return {"deleted": filename}


# ── Cohort builder ──────────────────────────────────────────────────────────

@app.get("/api/cohorts/dimensions")
def api_cohort_dimensions():
    """Aggregate cohort dimensions, value sets, and join info from all semantic profiles."""
    if not DATA_PROJECT or not PROFILE_BUCKET:
        raise HTTPException(503, "Project or bucket not configured")

    cache_key = f"{DATA_PROJECT}:{PROFILE_BUCKET}"
    now = _time.time()
    cached = _cohort_dims_cache.get(cache_key)
    if cached is not None:
        ts, resp = cached
        if now - ts < _SLOW_CACHE_TTL:
            return resp

    from concurrent.futures import ThreadPoolExecutor, as_completed
    from verily_profiler.storage import read_sem_profile, read_tech_profile

    try:
        avail = _cached_scan()
    except Exception:
        avail = {}
    sem_fqs = [fq for fq, info in avail.items() if info.get("semantic")]
    tech_fqs = [fq for fq, info in avail.items() if info.get("technical")]

    profiles: dict[str, dict | None] = {}
    tech_profiles: dict[str, dict | None] = {}

    with ThreadPoolExecutor(max_workers=16) as pool:
        futures = {}
        for fq in sem_fqs:
            futures[pool.submit(read_sem_profile, PROFILE_BUCKET, fq, project_id=BILLING_PROJECT)] = ("sem", fq)
        for fq in tech_fqs:
            futures[pool.submit(read_tech_profile, PROFILE_BUCKET, fq, project_id=BILLING_PROJECT)] = ("tech", fq)
        for f in as_completed(futures):
            kind, fq = futures[f]
            try:
                data = f.result()
            except Exception:
                data = None
            if kind == "sem":
                profiles[fq] = data
            else:
                tech_profiles[fq] = data

    import re as _re

    _ANCHOR_PATTERNS = [
        r"^usubjid$", r"^subjid$", r"^person_id$", r"^subject_id$", r"^patient_id$",
        r"^member_id$", r"^participant_id$", r"^enrolid$", r".*_id$", r"^id$",
    ]

    def _infer_anchor(sem: dict | None, col_names: list[str]) -> str:
        """Entity anchor for COUNT(DISTINCT): LLM value, else primary key, else an id-like column."""
        anchor = (sem or {}).get("entity_anchor") or ""
        if anchor:
            return anchor
        pk_cols = ((sem or {}).get("primary_key") or {}).get("columns") or []
        if len(pk_cols) == 1:
            return pk_cols[0]
        for pat in _ANCHOR_PATTERNS:
            for n in col_names:
                if _re.match(pat, n, _re.I):
                    return n
        return ""

    # Every column is a filterable dimension. We expose them all (LLM-picked
    # cohort_dimensions float to the top), enriched with data type + value sets
    # from whichever profile exists. A table needs only a technical OR semantic
    # profile to be cohortable.
    tables = []
    for fq in sorted(set(sem_fqs) | set(tech_fqs)):
        sem = profiles.get(fq)
        tech = tech_profiles.get(fq)
        sem_cols = (sem or {}).get("columns") or []
        tech_cols = (tech or {}).get("columns") or []

        tech_types = {c.get("name", ""): c.get("data_type", "STRING") for c in tech_cols}
        tech_values = {c.get("name", ""): (c.get("top_values") or []) for c in tech_cols}
        sem_by_name = {c.get("name", ""): c for c in sem_cols}

        col_names = [c.get("name", "") for c in (sem_cols or tech_cols) if c.get("name")]
        if not col_names:
            continue
        preferred = set((sem or {}).get("cohort_dimensions") or [])
        col_names.sort(key=lambda n: (n not in preferred, n.lower()))

        dims = []
        for name in col_names:
            sc = sem_by_name.get(name, {})
            dims.append({
                "column": name,
                "definition": sc.get("definition", ""),
                "values": sc.get("value_set_binding") or tech_values.get(name) or [],
                "data_type": tech_types.get(name, "STRING"),
                "preferred": name in preferred,
            })

        joinable = set()
        for col_data in sem_cols:
            for jp in col_data.get("join_paths") or []:
                parts = jp.rsplit(".", 1)
                if len(parts) == 2 and parts[0] != fq:
                    joinable.add(parts[0])

        tables.append({
            "fq_table": fq,
            "business_name": (sem or {}).get("business_name", "") or fq.rsplit(".", 1)[-1],
            "entity_anchor": _infer_anchor(sem, col_names),
            "entity_type": (sem or {}).get("entity_type", ""),
            "dimensions": dims,
            "joinable_tables": sorted(joinable),
        })

    tables.sort(key=lambda t: t["fq_table"])
    response = {"tables": tables}
    _cohort_dims_cache[cache_key] = (_time.time(), response)
    return response


_col_values_cache: dict[str, tuple[float, Any]] = {}


@app.get("/api/cohorts/column-values")
def api_column_values(table: str, column: str):
    """Return top values and data type for a column from its technical profile."""
    if not PROFILE_BUCKET:
        raise HTTPException(503, "Profile bucket not configured")

    cache_key = f"{table}:{column}"
    now = _time.time()
    cached = _col_values_cache.get(cache_key)
    if cached is not None:
        ts, resp = cached
        if now - ts < _SLOW_CACHE_TTL:
            return resp

    from verily_profiler.storage import read_tech_profile

    tech = read_tech_profile(PROFILE_BUCKET, table, project_id=BILLING_PROJECT)
    if not tech:
        return {"column": column, "data_type": "STRING", "values": []}

    for col in tech.get("columns", []):
        if col.get("name") == column:
            values = col.get("top_values") or []
            data_type = col.get("data_type", "STRING")
            response = {"column": column, "data_type": data_type, "values": values}
            _col_values_cache[cache_key] = (_time.time(), response)
            return response

    response = {"column": column, "data_type": "STRING", "values": []}
    _col_values_cache[cache_key] = (_time.time(), response)
    return response


ALLOWED_OPS = {"=", "!=", ">", ">=", "<", "<=", "LIKE", "IN", "NOT IN", "IS NULL", "IS NOT NULL", "BETWEEN"}
_IDENTIFIER_RE = None
_FQ_TABLE_RE = None

def _safe_fq_table(fq: str) -> tuple[str, str, str]:
    """Validate and parse a fully-qualified table name for safe SQL interpolation."""
    import re
    global _FQ_TABLE_RE
    if _FQ_TABLE_RE is None:
        _FQ_TABLE_RE = re.compile(r"^[a-zA-Z0-9_-]+\.[a-zA-Z0-9_]+\.[a-zA-Z0-9_]+$")
    if not _FQ_TABLE_RE.match(fq):
        raise HTTPException(400, f"Invalid table reference: {fq}")
    return parse_fq_table(fq)

def _safe_ident(name: str) -> str:
    import re
    global _IDENTIFIER_RE
    if _IDENTIFIER_RE is None:
        _IDENTIFIER_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
    if not _IDENTIFIER_RE.match(name):
        raise HTTPException(400, f"Invalid identifier: {name}")
    return name


def _parse_val(val: str):
    try:
        v = float(val)
        return int(v) if v == int(v) else v
    except (ValueError, TypeError):
        return val


def _build_where(filters: list[dict], alias: str) -> tuple[list[str], list[dict]]:
    clauses: list[str] = []
    params: list[dict] = []
    for i, f in enumerate(filters):
        col = _safe_ident(f["column"])
        op = f.get("operator", "=").strip()
        if op not in ALLOWED_OPS:
            raise HTTPException(400, f"Invalid operator: {op}")
        val = str(f.get("value", "")).strip()
        param_name = f"p_{alias}_{i}"

        if op in ("IS NULL", "IS NOT NULL"):
            clauses.append(f"{alias}.{col} {op}")
        elif op in ("IN", "NOT IN"):
            items = [v.strip() for v in val.split(",") if v.strip()]
            if not items:
                continue
            neg = "NOT " if op == "NOT IN" else ""
            clauses.append(f"{alias}.{col} {neg}IN UNNEST(@{param_name})")
            params.append({"name": param_name, "value": items, "type": "array"})
        elif op == "LIKE":
            clauses.append(f"{alias}.{col} LIKE @{param_name}")
            params.append({"name": param_name, "value": val, "type": "scalar"})
        elif op == "BETWEEN":
            parts = [v.strip() for v in val.split(",")]
            if len(parts) != 2:
                raise HTTPException(400, f"BETWEEN requires two comma-separated values, got: {val}")
            clauses.append(f"{alias}.{col} BETWEEN @{param_name}_lo AND @{param_name}_hi")
            params.append({"name": f"{param_name}_lo", "value": _parse_val(parts[0]), "type": "scalar"})
            params.append({"name": f"{param_name}_hi", "value": _parse_val(parts[1]), "type": "scalar"})
        else:
            clauses.append(f"{alias}.{col} {op} @{param_name}")
            params.append({"name": param_name, "value": _parse_val(val), "type": "scalar"})
    return clauses, params


def _make_bq_params(params: list[dict]):
    from google.cloud import bigquery as _bq
    bq_params = []
    for p in params:
        if p["type"] == "array":
            bq_params.append(_bq.ArrayQueryParameter(p["name"], "STRING", p["value"]))
        else:
            val = p["value"]
            if isinstance(val, float):
                bq_params.append(_bq.ScalarQueryParameter(p["name"], "FLOAT64", val))
            elif isinstance(val, int):
                bq_params.append(_bq.ScalarQueryParameter(p["name"], "INT64", val))
            else:
                bq_params.append(_bq.ScalarQueryParameter(p["name"], "STRING", str(val)))
    return bq_params


@app.post("/api/cohorts/execute")
def api_cohort_execute(body: dict[str, Any]):
    """Build SQL from cohort definition and execute on BigQuery."""
    if not BILLING_PROJECT:
        raise HTTPException(503, "Billing project not configured")
    from google.cloud import bigquery
    from bq_preview import _serialize_cell

    base = body.get("base_table", "")
    entity_col_raw = body.get("entity_column", "")
    entity_col = _safe_ident(entity_col_raw) if entity_col_raw else ""
    filters = body.get("filters") or []
    joins = body.get("joins") or []
    mode = body.get("mode", "count")

    p, d, t = _safe_fq_table(base)
    base_fq = f"`{p}.{d}.{t}`"

    all_clauses: list[str] = []
    all_params: list[dict] = []

    base_clauses, base_params = _build_where(filters, "t0")
    all_clauses.extend(base_clauses)
    all_params.extend(base_params)

    join_frags = []
    for ji, j in enumerate(joins):
        jt = j.get("target_table", "")
        jp, jd, jt_name = _safe_fq_table(jt)
        join_col = _safe_ident(j.get("join_column", entity_col or ""))
        alias = f"j{ji}"
        join_frags.append(
            f"JOIN `{jp}.{jd}.{jt_name}` {alias} ON t0.{join_col} = {alias}.{join_col}"
        )
        j_clauses, j_params = _build_where(j.get("filters") or [], alias)
        all_clauses.extend(j_clauses)
        all_params.extend(j_params)

    where = f"\nWHERE {' AND '.join(all_clauses)}" if all_clauses else ""
    joins_sql = "\n".join(join_frags)

    if mode == "preview":
        sql = f"SELECT t0.*\nFROM {base_fq} t0\n{joins_sql}{where}\nLIMIT 200"
    elif entity_col:
        sql = f"SELECT COUNT(DISTINCT t0.{entity_col}) AS cohort_count\nFROM {base_fq} t0\n{joins_sql}{where}"
    else:
        sql = f"SELECT COUNT(*) AS cohort_count\nFROM {base_fq} t0\n{joins_sql}{where}"

    client = bigquery.Client(project=BILLING_PROJECT)
    job_config = bigquery.QueryJobConfig(query_parameters=_make_bq_params(all_params)) if all_params else None

    try:
        job = client.query(sql, job_config=job_config)
        rows = list(job.result(timeout=120))
    except Exception as e:
        raise HTTPException(400, _friendly_query_error(e))

    if mode == "preview":
        schema = [{"name": f.name, "type": f.field_type} for f in (job.schema or [])]
        out_rows = [[_serialize_cell(r[k]) for k in r.keys()] for r in rows]
        return {"sql": sql, "columns": schema, "rows": out_rows, "row_count": len(out_rows)}
    else:
        count = rows[0]["cohort_count"] if rows else 0
        return {"sql": sql, "count": count}


def _friendly_query_error(e: Exception) -> str:
    msg = str(e)
    if "Not found: Table" in msg:
        return "Table not found. It may have been deleted or renamed."
    if "Access Denied" in msg or "403" in msg:
        return "Permission denied. Check your billing project has access to this table."
    if "Unrecognized name" in msg:
        return "Column not found. The selected dimension may no longer exist in this table."
    if "Could not cast" in msg or "Invalid" in msg:
        return "Value type mismatch. Check that filter values match the column data type."
    return f"Query error: {msg[:200]}"


# ── Cohort from terminology ──────────────────────────────────────────────────

@app.post("/api/cohorts/from-terminology")
def api_cohort_from_terminology(body: dict[str, Any]):
    """Build and execute a cohort query from terminology-based filters.

    Body: { filters: [{concept_key, fq_table, column, operator?, value?}, ...], mode }
    Filters without operator+value default to IS NOT NULL.
    Filters across multiple tables are joined via entity anchors.
    """
    if not BILLING_PROJECT:
        raise HTTPException(503, "Billing project not configured")
    if not DATA_PROJECT or not PROFILE_BUCKET:
        raise HTTPException(503, "Project or bucket not configured")

    from google.cloud import bigquery
    from bq_preview import _serialize_cell
    from verily_profiler.storage import read_sem_profile

    term_filters = body.get("filters") or []
    mode = body.get("mode", "count")
    if not term_filters:
        raise HTTPException(400, "filters list is required")

    table_filters: dict[str, list[dict]] = {}
    for f in term_filters:
        fq_table = f.get("fq_table", "")
        if not fq_table:
            continue
        table_filters.setdefault(fq_table, []).append(f)

    if not table_filters:
        raise HTTPException(400, "No valid filters provided")

    all_tables = list(table_filters.keys())
    base_table = all_tables[0]

    sem = read_sem_profile(PROFILE_BUCKET, base_table, project_id=BILLING_PROJECT)
    entity_col = (sem or {}).get("entity_anchor", "")

    if len(all_tables) > 1 and not entity_col:
        raise HTTPException(400, f"No entity anchor found for {base_table} — required for cross-table joins")
    if entity_col:
        _safe_ident(entity_col)

    pk_cols = ((sem or {}).get("primary_key") or {}).get("columns") or []

    p, d, t = _safe_fq_table(base_table)
    base_fq = f"`{p}.{d}.{t}`"

    all_clauses: list[str] = []
    all_params: list[dict] = []

    def _add_term_filters(filters: list[dict], alias: str):
        for i, f in enumerate(filters):
            col = _safe_ident(f.get("column", ""))
            op = f.get("operator", "").strip()
            val = f.get("value", "").strip() if f.get("value") else ""
            if not op or (not val and op not in ("IS NULL", "IS NOT NULL")):
                all_clauses.append(f"{alias}.{col} IS NOT NULL")
            else:
                fc, fp = _build_where([{"column": col, "operator": op, "value": val}], alias)
                for j, c in enumerate(fc):
                    c_renamed = c.replace(f"p_{alias}_0", f"p_{alias}_{i}_{j}")
                    all_clauses.append(c_renamed)
                for p in fp:
                    p["name"] = p["name"].replace(f"p_{alias}_0", f"p_{alias}_{i}_{j}")
                    all_params.append(p)

    _add_term_filters(table_filters[base_table], "t0")

    join_frags = []
    for ji, jt in enumerate(all_tables[1:]):
        jp, jd, jt_name = _safe_fq_table(jt)
        alias = f"j{ji}"
        join_frags.append(
            f"JOIN `{jp}.{jd}.{jt_name}` {alias} ON t0.{entity_col} = {alias}.{entity_col}"
        )
        _add_term_filters(table_filters[jt], alias)

    where = f"\nWHERE {' AND '.join(all_clauses)}" if all_clauses else ""
    joins_sql = "\n".join(join_frags)

    if mode == "preview":
        sql = f"SELECT t0.*\nFROM {base_fq} t0\n{joins_sql}{where}\nLIMIT 200"
    elif entity_col:
        sql = f"SELECT COUNT(DISTINCT t0.{entity_col}) AS cohort_count\nFROM {base_fq} t0\n{joins_sql}{where}"
    elif pk_cols:
        pk_expr = ", ".join(f"t0.{_safe_ident(c)}" for c in pk_cols)
        sql = f"SELECT COUNT(DISTINCT CONCAT({pk_expr})) AS cohort_count\nFROM {base_fq} t0{where}"
    else:
        sql = f"SELECT COUNT(*) AS cohort_count\nFROM {base_fq} t0{where}"

    client = bigquery.Client(project=BILLING_PROJECT)
    job_config = bigquery.QueryJobConfig(query_parameters=_make_bq_params(all_params)) if all_params else None

    try:
        job = client.query(sql, job_config=job_config)
        rows = list(job.result(timeout=120))
    except Exception as e:
        raise HTTPException(400, _friendly_query_error(e))

    tables_used = all_tables
    if mode == "preview":
        schema = [{"name": f.name, "type": f.field_type} for f in (job.schema or [])]
        out_rows = [[_serialize_cell(r[k]) for k in r.keys()] for r in rows]
        return {
            "sql": sql,
            "columns": schema,
            "rows": out_rows,
            "row_count": len(out_rows),
            "base_table": base_table,
            "tables_used": tables_used,
        }
    else:
        count = rows[0]["cohort_count"] if rows else 0
        return {
            "sql": sql,
            "count": count,
            "base_table": base_table,
            "tables_used": tables_used,
        }


# ── Cohort from natural language ─────────────────────────────────────────────

@app.post("/api/cohorts/from-natural-language")
async def api_cohort_from_nl(body: dict[str, Any]):
    """Generate and optionally execute a cohort SQL query from a natural language description."""
    if not BILLING_PROJECT:
        raise HTTPException(503, "Billing project not configured")
    if not DATA_PROJECT or not PROFILE_BUCKET:
        raise HTTPException(503, "Project or bucket not configured")

    query = (body.get("query") or "").strip()
    mode = body.get("mode", "generate")
    if not query:
        raise HTTPException(400, "query is required")

    import asyncio
    from chat_handler import build_context, _load_catalog_context_md

    catalog_md = _load_catalog_context_md(DATA_PROJECT, PROFILE_BUCKET, BILLING_PROJECT)
    if not catalog_md:
        raise HTTPException(400, "No catalog context available. Profile tables first.")

    model = GEMINI_MODEL or detect_available_model(BILLING_PROJECT)
    if not model:
        raise HTTPException(503, "No Gemini model available")

    system_prompt = f"""You are a cohort query builder for BigQuery datasets.

Given a natural language description, generate a BigQuery SQL query that identifies a cohort of subjects.

RULES:
- Use only tables and columns from the catalog context below.
- Always use fully-qualified table names (`project.dataset.table`).
- For count queries, use COUNT(DISTINCT entity_column) AS cohort_count.
- For preview queries, SELECT * with LIMIT 200.
- Return ONLY valid BigQuery SQL, no markdown fences.
- After the SQL, on a new line starting with "EXPLANATION:", provide a 1-2 sentence explanation.

CATALOG CONTEXT:
{catalog_md}"""

    from google.cloud import aiplatform
    import vertexai
    from vertexai.generative_models import GenerativeModel

    from verily_profiler.llm import _get_location
    vertexai.init(project=BILLING_PROJECT, location=_get_location())
    gen_model = GenerativeModel(model)

    user_prompt = f"Generate a cohort COUNT query for: {query}"

    def _generate():
        response = gen_model.generate_content(
            [system_prompt, user_prompt],
            generation_config={"temperature": 0.1, "max_output_tokens": 2048},
        )
        return response.text

    raw = await asyncio.to_thread(_generate)

    lines = raw.strip().split("\n")
    sql_lines = []
    explanation = ""
    for line in lines:
        if line.strip().startswith("EXPLANATION:"):
            explanation = line.strip().removeprefix("EXPLANATION:").strip()
        elif line.strip().startswith("```"):
            continue
        else:
            sql_lines.append(line)
    sql = "\n".join(sql_lines).strip()

    result: dict[str, Any] = {"sql": sql, "explanation": explanation}

    if mode == "execute" and sql:
        from google.cloud import bigquery
        from bq_preview import _serialize_cell

        client = bigquery.Client(project=BILLING_PROJECT)
        try:
            job = client.query(sql)
            rows = list(job.result(timeout=120))
        except Exception as e:
            result["error"] = f"Query failed: {e}"
            return result

        schema = [{"name": f.name, "type": f.field_type} for f in (job.schema or [])]
        out_rows = [[_serialize_cell(r[k]) for k in r.keys()] for r in rows]

        is_count = False
        count_val = None
        if len(rows) == 1:
            keys = list(rows[0].keys())
            if len(keys) == 1:
                val = rows[0][keys[0]]
                if isinstance(val, (int, float)):
                    is_count = True
                    count_val = int(val)

        if is_count:
            result["count"] = count_val
            import re
            preview = re.sub(
                r"(?is)^SELECT\s+COUNT\s*\(.*?\)\s+AS\s+\w+",
                "SELECT *",
                sql,
            )
            if "LIMIT" not in preview.upper():
                preview += "\nLIMIT 200"
            result["preview_sql"] = preview
        else:
            result["columns"] = schema
            result["rows"] = out_rows
            result["row_count"] = len(out_rows)

    return result


@app.post("/api/cohorts/run-preview")
def api_run_preview(body: dict[str, Any]):
    """Execute a read-only preview SQL (must start with SELECT, must contain LIMIT)."""
    if not BILLING_PROJECT:
        raise HTTPException(503, "Billing project not configured")

    sql = (body.get("sql") or "").strip()
    if not sql:
        raise HTTPException(400, "sql is required")
    normalized = sql.lstrip().upper()
    if not normalized.startswith("SELECT"):
        raise HTTPException(400, "Only SELECT queries are allowed")
    if "LIMIT" not in normalized:
        raise HTTPException(400, "Query must contain a LIMIT clause")

    from google.cloud import bigquery
    from bq_preview import _serialize_cell

    client = bigquery.Client(project=BILLING_PROJECT)
    try:
        job = client.query(sql)
        rows = list(job.result(timeout=120))
    except Exception as e:
        raise HTTPException(400, f"Query failed: {e}")

    schema = [{"name": f.name, "type": f.field_type} for f in (job.schema or [])]
    out_rows = [[_serialize_cell(r[k]) for k in r.keys()] for r in rows]
    return {"sql": sql, "columns": schema, "rows": out_rows, "row_count": len(out_rows)}


# ── Reset profiles ──────────────────────────────────────────────────────────

@app.delete("/api/profiles/reset")
def api_reset_profiles():
    """Delete ALL profiling artifacts from GCS for the current data project."""
    if not DATA_PROJECT or not PROFILE_BUCKET:
        raise HTTPException(503, "Project or bucket not configured")
    from google.cloud import storage

    client = storage.Client(project=BILLING_PROJECT)
    bucket = client.bucket(PROFILE_BUCKET.replace("gs://", ""))
    prefix = f"profiling/{DATA_PROJECT}/"
    blobs = list(bucket.list_blobs(prefix=prefix))

    artifacts = [f"_terminology_registry_{DATA_PROJECT}.json", f"_catalog_context_{DATA_PROJECT}.md"]
    for name in artifacts:
        b = bucket.blob(name)
        if b.exists():
            blobs.append(b)

    deleted = 0
    for blob in blobs:
        try:
            blob.delete()
            deleted += 1
        except Exception:
            pass

    _invalidate_profiling_caches()
    from chat_handler import invalidate_context_cache
    invalidate_context_cache(DATA_PROJECT, PROFILE_BUCKET)

    logger.info(f"Reset profiles: deleted {deleted} objects for {DATA_PROJECT}")
    return {"deleted_count": deleted, "data_project": DATA_PROJECT}


# ── Catalog context regeneration ─────────────────────────────────────────────

@app.post("/api/catalog-context/regenerate")
def api_regenerate_context():
    """Manually regenerate the catalog context .md file."""
    if not DATA_PROJECT or not PROFILE_BUCKET:
        raise HTTPException(503, "Project or bucket not configured")
    from verily_profiler.storage import regenerate_catalog_context
    from chat_handler import invalidate_context_cache
    try:
        path = regenerate_catalog_context(PROFILE_BUCKET, DATA_PROJECT, billing_project_id=BILLING_PROJECT)
        invalidate_context_cache(DATA_PROJECT, PROFILE_BUCKET)
        return {"status": "ok", "path": path}
    except Exception as e:
        raise HTTPException(500, f"Regeneration failed: {e}")


# ── Terminology registry ────────────────────────────────────────────────────

_terminology_slim_cache: dict[str, tuple[float, Any]] = {}


@app.get("/api/terminology/slim")
def api_terminology_slim():
    """Lightweight terminology list — reads only the registry, no profile loading."""
    if not DATA_PROJECT or not PROFILE_BUCKET:
        raise HTTPException(503, "Project or bucket not configured")

    cache_key = f"{DATA_PROJECT}:{PROFILE_BUCKET}"
    now = _time.time()
    cached = _terminology_slim_cache.get(cache_key)
    if cached is not None:
        ts, resp = cached
        if now - ts < _SLOW_CACHE_TTL:
            return resp

    from verily_profiler.storage import read_registry

    registry = read_registry(PROFILE_BUCKET, DATA_PROJECT, BILLING_PROJECT)
    entries = []
    for e in registry.entries:
        tables = set()
        for sc in e.source_columns:
            parts = sc.rsplit(".", 1)
            if len(parts) == 2:
                tables.add(parts[0])
        entries.append({
            "system": e.system,
            "code": e.code,
            "display": e.display,
            "concept_key": e.concept_key,
            "source_columns": e.source_columns,
            "tables_count": len(tables),
            "columns_count": len(e.source_columns),
        })
    response = {
        "entries": entries,
        "total": len(entries),
        "updated_at": registry.updated_at,
    }
    _terminology_slim_cache[cache_key] = (_time.time(), response)
    return response


@app.get("/api/terminology")
def api_terminology():
    """Return the cross-table terminology registry with per-entry counts and column metadata."""
    if not DATA_PROJECT or not PROFILE_BUCKET:
        raise HTTPException(503, "Project or bucket not configured")

    cache_key = f"{DATA_PROJECT}:{PROFILE_BUCKET}"
    now = _time.time()
    cached = _terminology_cache.get(cache_key)
    if cached is not None:
        ts, resp = cached
        if now - ts < _SLOW_CACHE_TTL:
            return resp

    from concurrent.futures import ThreadPoolExecutor
    from verily_profiler.storage import read_registry, read_sem_profile

    registry = read_registry(PROFILE_BUCKET, DATA_PROJECT, BILLING_PROJECT)

    needed_tables = set()
    for e in registry.entries:
        for sc in e.source_columns:
            parts = sc.rsplit(".", 1)
            if len(parts) == 2:
                needed_tables.add(parts[0])

    sem_local: dict[str, dict | None] = {}

    def _load(fq: str):
        return fq, read_sem_profile(PROFILE_BUCKET, fq, project_id=BILLING_PROJECT)

    with ThreadPoolExecutor(max_workers=10) as pool:
        for fq, data in pool.map(lambda f: _load(f), needed_tables):
            sem_local[fq] = data

    def _get_sem(fq_table: str) -> dict | None:
        return sem_local.get(fq_table)

    def _col_meta(fq_col: str) -> dict:
        parts = fq_col.rsplit(".", 1)
        if len(parts) != 2:
            return {"fq_column": fq_col}
        fq_table, col_name = parts
        result: dict = {"fq_column": fq_col, "column": col_name, "fq_table": fq_table}
        sem = _get_sem(fq_table)
        if sem:
            for c in sem.get("columns", []):
                if c.get("name") == col_name:
                    result["definition"] = c.get("definition", "")
                    result["measurement_method"] = c.get("measurement_method", "")
                    break
        return result

    entries = []
    for e in registry.entries:
        tables = set()
        columns_meta = []
        for sc in e.source_columns:
            meta = _col_meta(sc)
            columns_meta.append(meta)
            if "fq_table" in meta:
                tables.add(meta["fq_table"])
        entries.append({
            "system": e.system,
            "code": e.code,
            "display": e.display,
            "concept_key": e.concept_key,
            "source_columns": e.source_columns,
            "columns_meta": columns_meta,
            "tables_count": len(tables),
            "columns_count": len(e.source_columns),
        })
    response = {
        "entries": entries,
        "total": len(entries),
        "updated_at": registry.updated_at,
    }
    _terminology_cache[cache_key] = (_time.time(), response)
    return response


# ── Chat ─────────────────────────────────────────────────────────────────────

from chat_handler import chat_store, handle_chat_message, handle_chat_stream
from starlette.responses import StreamingResponse


@app.post("/api/chat")
async def api_chat(body: dict[str, Any]):
    """
    Send a chat message.
    Body: { message, mode?, fq_table?, session_id? }
    Returns: { session_id, message: ChatMessage }
    """
    message = body.get("message", "").strip()
    if not message:
        raise HTTPException(400, "message is required")
    if not DATA_PROJECT:
        raise HTTPException(503, "No data project configured")

    mode = body.get("mode", "metadata")
    if mode not in ("metadata", "agent"):
        mode = "metadata"

    fq_table = body.get("fq_table") or None
    session_id = body.get("session_id") or None
    detail_level = body.get("detail_level", "summary")
    if detail_level not in ("summary", "full"):
        detail_level = "summary"

    try:
        result = await handle_chat_message(
            message=message,
            mode=mode,
            fq_table=fq_table,
            session_id=session_id,
            data_project=DATA_PROJECT,
            billing_project=BILLING_PROJECT,
            bucket=PROFILE_BUCKET,
            model=CHAT_MODEL,
            detail_level=detail_level,
        )
        return result
    except Exception as e:
        import traceback
        logger.error(f"Chat error:\n{traceback.format_exc()}")
        raise HTTPException(500, f"Chat failed: {e}")


@app.post("/api/chat/stream")
async def api_chat_stream(body: dict[str, Any]):
    """SSE streaming chat — prevents 504 on long LLM calls."""
    message = body.get("message", "").strip()
    if not message:
        raise HTTPException(400, "message is required")
    if not DATA_PROJECT:
        raise HTTPException(503, "No data project configured")

    mode = body.get("mode", "metadata")
    if mode not in ("metadata", "agent"):
        mode = "metadata"

    fq_table = body.get("fq_table") or None
    session_id = body.get("session_id") or None
    detail_level = body.get("detail_level", "summary")
    if detail_level not in ("summary", "full"):
        detail_level = "summary"

    async def event_generator():
        async for event in handle_chat_stream(
            message=message,
            mode=mode,
            fq_table=fq_table,
            session_id=session_id,
            data_project=DATA_PROJECT,
            billing_project=BILLING_PROJECT,
            bucket=PROFILE_BUCKET,
            model=CHAT_MODEL,
            detail_level=detail_level,
        ):
            yield event

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@app.post("/api/chat/clear")
def api_chat_clear(body: dict[str, Any]):
    """Clear a chat session. Body: { session_id }"""
    sid = body.get("session_id", "")
    if not sid:
        raise HTTPException(400, "session_id is required")
    chat_store.clear(sid)
    return {"status": "cleared", "session_id": sid}


@app.get("/api/chat/history/{session_id}")
def api_chat_history(session_id: str):
    """Get conversation history for a session."""
    sess = chat_store.get(session_id)
    if not sess:
        raise HTTPException(404, "Session not found")
    return {
        "session_id": sess.session_id,
        "mode": sess.mode,
        "messages": [m.to_json_dict() for m in sess.messages],
    }


@app.post("/api/chat/preload")
async def api_chat_preload():
    """Pre-load all profiles into the in-memory cache so the next chat message is fast."""
    if not DATA_PROJECT or not PROFILE_BUCKET:
        raise HTTPException(503, "Project or bucket not configured")
    import asyncio
    from chat_handler import _load_all_profiles
    await asyncio.to_thread(_load_all_profiles, PROFILE_BUCKET, DATA_PROJECT, BILLING_PROJECT)
    return {"status": "ok"}


@app.get("/api/chat/context-info")
def api_chat_context_info():
    """Report what context is available for grounding the chat."""
    if not DATA_PROJECT or not PROFILE_BUCKET:
        return {"status": "unconfigured", "profiled_tables": 0, "catalog_context_chars": 0}
    from chat_handler import _load_catalog_context_md
    catalog_md = _load_catalog_context_md(DATA_PROJECT, PROFILE_BUCKET, BILLING_PROJECT)
    try:
        avail = _cached_scan()
    except Exception:
        avail = {}
    profiled = [fq for fq, info in avail.items() if info.get("technical") or info.get("semantic")]
    fully_profiled = [fq for fq, info in avail.items() if info.get("technical") and info.get("semantic")]
    return {
        "status": "ready" if catalog_md else "no_context",
        "profiled_tables": len(profiled),
        "fully_profiled_tables": len(fully_profiled),
        "catalog_context_chars": len(catalog_md) if catalog_md else 0,
        "tables": [{"fq_table": fq, "has_tech": avail[fq].get("technical", False), "has_sem": avail[fq].get("semantic", False)} for fq in sorted(profiled)],
    }


# ── Static frontend (production build) ───────────────────────────────────────

if FRONTEND_DIST.is_dir():
    if (FRONTEND_DIST / "assets").is_dir():
        app.mount("/assets", StaticFiles(directory=FRONTEND_DIST / "assets"), name="assets")

    @app.get("/logo.png")
    def _serve_logo():
        logo = FRONTEND_DIST / "logo.png"
        if logo.is_file():
            return FileResponse(logo, media_type="image/png")
        raise HTTPException(404)

    @app.get("/{path:path}")
    def _spa_fallback(path: str):
        # Serve static file if it exists at root level
        candidate = FRONTEND_DIST / path
        if path and candidate.is_file() and ".." not in path:
            return FileResponse(candidate)
        # Otherwise serve index.html for SPA routing
        index = FRONTEND_DIST / "index.html"
        if index.is_file():
            return FileResponse(index, media_type="text/html")
        raise HTTPException(404)


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8080, reload=True)
